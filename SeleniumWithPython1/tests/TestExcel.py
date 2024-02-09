import unittest
import openpyxl



class MyTestCase(unittest.TestCase):
    def test_something(self):
        wb=openpyxl.load_workbook('Book1.xlsx')
        sheet1=wb['Sheet1']
        #print(sheet1.cell(row=3,column=2).value)
        #print(sheet1['B4'].value)

        #self.assertEqual(True, False)
        for x in range(2,5):
            print('ROW NO :',x,sheet1.cell(row=x,column=1).value,sheet1.cell(row=x,column=2).value)
            sheet1.cell(row=x,column=3).value='demoresult'
        wb.save('Book1.xlsx')
        """
        cells=sheet1['A2':'C4']
        for c1,c2,c3 in cells:
            print(c1.value,c2.value)
            c3.value='result';
        wb.save('Book1.xlsx')
        """

if __name__ == '__main__':
    unittest.main()
